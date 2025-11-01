from pathlib import Path
import re
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === CONFIG ===
pm_dir = Path("common/production_methods")  # folder containing all PM .txt files
goods_file = Path("common/goods/00_goods.txt")
output_file = Path("output/productivity_report.xlsx")
log_file = Path("output/parser_log.txt")
DEFAULT_EMPLOYEES = 10000

log_file.parent.mkdir(parents=True, exist_ok=True)
Path(output_file.parent).mkdir(parents=True, exist_ok=True)
log_lines = []

def log(msg):
    print(msg)
    log_lines.append(msg)

# === LOAD GOODS PRICES ===
def load_goods_prices(path: Path):
    prices = {}
    if not path.exists():
        log(f"[WARN] Goods file not found: {path}")
        return prices
    with open(path, encoding="utf-8", errors="ignore") as f:
        text = f.read()
    for m in re.finditer(r'\b([A-Za-z0-9_]+)\s*=\s*{', text):
        name = m.group(1)
        block_start = m.end()
        brace_count = 1
        i = block_start
        while brace_count > 0 and i < len(text):
            if text[i] == '{':
                brace_count += 1
            elif text[i] == '}':
                brace_count -= 1
            i += 1
        block_text = text[block_start:i-1]
        cost_match = re.search(r'\b(?:cost|base_price)\s*=\s*([0-9]+(?:\.[0-9]+)?)', block_text)
        if cost_match:
            prices[name] = float(cost_match.group(1))
    log(f"[INFO] Loaded {len(prices)} goods with prices.")
    return prices

# === PARSE SINGLE PM FILE ===
def parse_pm_file(path: Path, goods_prices):
    with open(path, encoding="utf-8", errors="ignore") as f:
        text = f.read()

    rows = []

    # Only match top-level PMs starting with "pm_"
    pm_pattern = re.compile(r'\b(pm_[a-zA-Z0-9_]+)\s*=\s*{')
    for match in pm_pattern.finditer(text):
        pm_name = match.group(1)
        start_idx = match.end()
        brace_count = 1
        i = start_idx

        total_inputs = {}
        total_outputs = {}
        total_employees = 0
        notes = []

        # Loop until the end of this PM block
        while brace_count > 0 and i < len(text):
            char = text[i]
            if char == '{':
                brace_count += 1
            elif char == '}':
                brace_count -= 1
            i += 1

        block_text = text[start_idx:i-1]

        # Collect all goods_input/output and employment inside PM
        for line in block_text.splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue

            m_in = re.match(r'goods_input_([A-Za-z0-9_]+)_add\s*=\s*([0-9]+(?:\.[0-9]+)?)', line)
            if m_in:
                g, amt = m_in.group(1), float(m_in.group(2))
                total_inputs[g] = total_inputs.get(g, 0.0) + amt

            m_out = re.match(r'goods_output_([A-Za-z0-9_]+)_add\s*=\s*([0-9]+(?:\.[0-9]+)?)', line)
            if m_out:
                g, amt = m_out.group(1), float(m_out.group(2))
                total_outputs[g] = total_outputs.get(g, 0.0) + amt

            m_emp = re.match(r'building_employment_[A-Za-z0-9_]+_add\s*=\s*([0-9]+(?:\.[0-9]+)?)', line)
            if m_emp:
                total_employees += float(m_emp.group(1))

        if total_employees == 0:
            total_employees = DEFAULT_EMPLOYEES
            notes.append(f"emp_fallback={DEFAULT_EMPLOYEES}")

        total_in_val, total_out_val, missing_goods = 0.0, 0.0, set()
        for g, amt in total_inputs.items():
            if g in goods_prices:
                total_in_val += amt * goods_prices[g]
            else:
                missing_goods.add(g)
        for g, amt in total_outputs.items():
            if g in goods_prices:
                total_out_val += amt * goods_prices[g]
            else:
                missing_goods.add(g)
        if missing_goods:
            notes.append("missing_prices=" + ";".join(sorted(missing_goods)))

        total_profit = total_out_val - total_in_val
        productivity = (total_profit / total_employees) * 100  # x100 for visualization

        # Skip automation/placeholder PMs: no output or negative/zero employees
        if total_out_val == 0 or total_employees <= 0:
            log(f"[SKIP] PM '{pm_name}' skipped (automation/placeholder)")
            continue

        rows.append({
            "BlockName": pm_name,
            "Employees": int(total_employees),
            "TotalInputValue": round(total_in_val,3),
            "TotalOutputValue": round(total_out_val,3),
            "TotalProfit": round(total_profit,3),
            "Productivity_x100": round(productivity,3),
            "Notes": ";".join(notes)
        })

        log(f"[DEBUG] PM '{pm_name}' → Profit: {round(total_profit,3)}, Employees: {int(total_employees)}")

    return rows

# === PARSE ALL PM FILES IN FOLDER ===
def parse_all_pm_files(pm_dir: Path, goods_prices):
    all_rows = []
    for file_path in pm_dir.glob("*.txt"):
        log(f"[INFO] Parsing PM file: {file_path.name}")
        rows = parse_pm_file(file_path, goods_prices)
        all_rows.extend(rows)
    return all_rows

# === WRITE EXCEL ===
def write_excel(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productivity Report"

    headers = ["BlockName","Employees","TotalInputValue","TotalOutputValue","TotalProfit","Productivity_x100","Notes"]
    ws.append(headers)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in rows:
        ws.append([row[h] for h in headers])
        last_row_idx = ws.max_row
        prod_cell = ws.cell(row=last_row_idx, column=headers.index("Productivity_x100")+1)
        if prod_cell.value >= 0:
            prod_cell.fill = green_fill
        else:
            prod_cell.fill = red_fill

    # Auto column widths
    for i, _ in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_path)
    log(f"[DONE] Excel written → {output_path.resolve()}")

# === MAIN ===
def main():
    goods_prices = load_goods_prices(goods_file)
    all_rows = parse_all_pm_files(pm_dir, goods_prices)
    if not all_rows:
        log("[WARN] No PMs were processed! Check paths and file contents.")
    write_excel(all_rows, output_file)

    # Write log file
    with open(log_file, "w", encoding="utf-8") as lf:
        lf.write("\n".join(log_lines))
    log(f"[INFO] Full log written → {log_file.resolve()}")
    input("Press Enter to exit...")

if __name__ == "__main__":
    main()
